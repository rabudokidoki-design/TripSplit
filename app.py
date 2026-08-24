from flask import Flask, send_from_directory, request, jsonify, send_file
import os, json, io, datetime, uuid, hashlib
from functools import wraps

app = Flask(__name__)
DATA_FILE = 'trips_data.json'

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE,'r') as f:
                data=json.load(f)
                # migrate old format
                if 'trips' not in data:
                    # old single-trip format
                    return {"trips": {}}
                return data
        except:
            return {"trips": {}}
    return {"trips": {}}

def save_data(data):
    with open(DATA_FILE,'w') as f:
        json.dump(data,f,indent=2)

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def require_trip_auth(f):
    @wraps(f)
    def decorated(trip_id, *args, **kwargs):
        data=load_data()
        trip=data['trips'].get(trip_id)
        if not trip:
            return jsonify({"error":"Trip not found"}),404
        # password can be in header X-Trip-Password or json body or query
        pw = request.headers.get('X-Trip-Password') or request.args.get('password') or (request.get_json(silent=True) or {}).get('password')
        if not pw:
            return jsonify({"error":"Password required"}),401
        if hash_pw(pw)!=trip.get('passwordHash'):
            return jsonify({"error":"Wrong password"}),403
        return f(trip_id, *args, **kwargs)
    return decorated

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    # prevent serving data file
    if path==DATA_FILE:
        return "",404
    return send_from_directory('.', path)

@app.route('/api/trips', methods=['GET'])
def list_trips():
    # For privacy, don't list all trips publicly unless ?admin=1
    # Return only count, or return list without sensitive data if requested with myTrips param
    data=load_data()
    # Only return id, name, friendCount, createdAt, expenseCount
    trips=[]
    for tid,t in data['trips'].items():
        trips.append({
            "id": tid,
            "name": t.get('name',''),
            "friendCount": len(t.get('friends',[])),
            "expenseCount": len(t.get('expenses',[])),
            "createdAt": t.get('createdAt',''),
            "total": sum([e.get('amount',0) for e in t.get('expenses',[])])
        })
    return jsonify({"trips": trips})

@app.route('/api/trips', methods=['POST'])
def create_trip():
    payload=request.get_json()
    name=payload.get('name','').strip()
    code=payload.get('code','').strip() or str(uuid.uuid4())[:6].upper()
    password=payload.get('password','').strip()
    friends=payload.get('friends',[])
    if not name:
        return jsonify({"error":"Trip name required"}),400
    if not password or len(password)<3:
        return jsonify({"error":"Password min 3 chars"}),400
    if not code:
        code=str(uuid.uuid4())[:6].upper()
    # sanitize code
    code=''.join([c for c in code if c.isalnum()])[:20].upper()
    if not code:
        code=str(uuid.uuid4())[:6].upper()
    data=load_data()
    if code in data['trips']:
        return jsonify({"error":f"Trip code {code} already exists, pick another"}),400
    trip={
        "id": code,
        "name": name,
        "passwordHash": hash_pw(password),
        "friends": [f for f in friends if f][:8],
        "expenses": [],
        "createdAt": datetime.datetime.now().isoformat(),
    }
    data['trips'][code]=trip
    save_data(data)
    # return without hash
    return jsonify({"id": code, "name": name, "friends": trip['friends']})

@app.route('/api/trips/<trip_id>/auth', methods=['POST'])
def auth_trip(trip_id):
    payload=request.get_json() or {}
    pw=payload.get('password','')
    data=load_data()
    trip=data['trips'].get(trip_id)
    if not trip:
        return jsonify({"error":"Trip not found"}),404
    if hash_pw(pw)!=trip.get('passwordHash'):
        return jsonify({"error":"Wrong password"}),403
    # return trip data (without hash)
    return jsonify({
        "id": trip['id'],
        "name": trip['name'],
        "friends": trip['friends'],
        "expenses": trip['expenses'],
        "createdAt": trip['createdAt']
    })

@app.route('/api/trips/<trip_id>', methods=['GET'])
@require_trip_auth
def get_trip(trip_id):
    data=load_data()
    trip=data['trips'][trip_id]
    return jsonify({
        "id": trip['id'],
        "name": trip['name'],
        "friends": trip['friends'],
        "expenses": trip['expenses'],
        "createdAt": trip['createdAt']
    })

@app.route('/api/trips/<trip_id>', methods=['PUT'])
@require_trip_auth
def update_trip(trip_id):
    payload=request.get_json() or {}
    data=load_data()
    trip=data['trips'][trip_id]
    if 'name' in payload:
        trip['name']=payload['name'][:50]
    if 'friends' in payload:
        trip['friends']=[f for f in payload['friends'] if f][:8]
        # clean expenses that reference removed friends? Keep but filter sharedWith
    save_data(data)
    return jsonify({"ok":True})

@app.route('/api/trips/<trip_id>', methods=['DELETE'])
@require_trip_auth
def delete_trip(trip_id):
    data=load_data()
    if trip_id in data['trips']:
        del data['trips'][trip_id]
        save_data(data)
    return jsonify({"ok":True})

@app.route('/api/trips/<trip_id>/expenses', methods=['POST'])
@require_trip_auth
def add_expense(trip_id):
    payload=request.get_json()
    data=load_data()
    trip=data['trips'][trip_id]
    exp={
        "id": payload.get('id') or int(datetime.datetime.now().timestamp()*1000),
        "date": payload.get('date') or datetime.date.today().isoformat(),
        "desc": payload.get('desc','Expense')[:100],
        "category": payload.get('category','Other'),
        "paidBy": payload.get('paidBy',''),
        "amount": float(payload.get('amount',0)),
        "sharedWith": payload.get('sharedWith',[]),
        "createdAt": datetime.datetime.now().isoformat()
    }
    if exp['paidBy'] not in trip['friends']:
        return jsonify({"error":"Payer not in friends"}),400
    if not exp['sharedWith']:
        return jsonify({"error":"Select at least 1 friend to split"}),400
    trip['expenses'].insert(0,exp)
    save_data(data)
    return jsonify(exp)

@app.route('/api/trips/<trip_id>/expenses/<int:exp_id>', methods=['PUT'])
@require_trip_auth
def edit_expense(trip_id, exp_id):
    payload=request.get_json()
    data=load_data()
    trip=data['trips'][trip_id]
    for i,ex in enumerate(trip['expenses']):
        if ex['id']==exp_id:
            # update fields
            if 'date' in payload: trip['expenses'][i]['date']=payload['date']
            if 'desc' in payload: trip['expenses'][i]['desc']=payload['desc'][:100]
            if 'category' in payload: trip['expenses'][i]['category']=payload['category']
            if 'paidBy' in payload: trip['expenses'][i]['paidBy']=payload['paidBy']
            if 'amount' in payload: trip['expenses'][i]['amount']=float(payload['amount'])
            if 'sharedWith' in payload: trip['expenses'][i]['sharedWith']=payload['sharedWith']
            save_data(data)
            return jsonify(trip['expenses'][i])
    return jsonify({"error":"Expense not found"}),404

@app.route('/api/trips/<trip_id>/expenses/<int:exp_id>', methods=['DELETE'])
@require_trip_auth
def delete_expense(trip_id, exp_id):
    data=load_data()
    trip=data['trips'][trip_id]
    trip['expenses']=[e for e in trip['expenses'] if e['id']!=exp_id]
    save_data(data)
    return jsonify({"ok":True})

@app.route('/api/trips/<trip_id>/export', methods=['GET'])
@require_trip_auth
def export_trip(trip_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    data=load_data()
    trip=data['trips'][trip_id]
    friends=trip['friends'][:8]
    while len(friends)<8:
        friends.append("")
    expenses=trip['expenses']

    wb=openpyxl.Workbook()
    header_fill=PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font=Font(color="FFFFFF", bold=True, size=11)
    sub_fill=PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    green_fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill=PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    money_fmt='#,##0.00'

    ws_f=wb.active
    ws_f.title="2. Friends"
    ws_f['A1']=f"Trip: {trip['name']} ({trip['id']})"
    ws_f['A1'].font=Font(bold=True,size=14,color="2F5597")
    for c,h in enumerate(["No.","Name"], start=1):
        cell=ws_f.cell(row=3,column=c,value=h)
        cell.fill=header_fill
        cell.font=header_font
    for i,name in enumerate(friends):
        if name:
            ws_f.cell(row=4+i,column=1,value=i+1)
            ws_f.cell(row=4+i,column=2,value=name)
    from openpyxl.workbook.defined_name import DefinedName
    dn=DefinedName(name="FriendsList", attr_text="'2. Friends'!$B$4:$B$11")
    wb.defined_names[dn.name]=dn

    ws_e=wb.create_sheet("3. Expenses - Mobile")
    ws_e.freeze_panes="G5"
    headers=[(1,"ID"),(2,"Date"),(3,"What"),(4,"Category"),(5,"Paid By"),(6,"Amount"),(7,"# Sharing")]
    for col,title in headers:
        c=ws_e.cell(row=4,column=col,value=title)
        c.fill=header_fill
        c.font=header_font
    for idx,name in enumerate(friends):
        if not name: continue
        col=8+idx
        ws_e.cell(row=3,column=col,value=name).font=Font(bold=True)
        ws_e.cell(row=3,column=col).fill=sub_fill
        ws_e.cell(row=4,column=col,value="Y/N").fill=header_fill
        ws_e.cell(row=4,column=col).font=header_font
    for r,exp in enumerate(expenses, start=5):
        ws_e.cell(row=r,column=1,value=r-4)
        ws_e.cell(row=r,column=2,value=exp.get('date',''))
        ws_e.cell(row=r,column=3,value=exp.get('desc',''))
        ws_e.cell(row=r,column=4,value=exp.get('category',''))
        ws_e.cell(row=r,column=5,value=exp.get('paidBy',''))
        ws_e.cell(row=r,column=6,value=float(exp.get('amount',0))).number_format=money_fmt
        shared=exp.get('sharedWith',[])
        ws_e.cell(row=r,column=7,value=len(shared))
        for idx,name in enumerate(friends):
            if not name: continue
            col=8+idx
            ws_e.cell(row=r,column=col,value="Y" if name in shared else "N")
    for idx in range(len([f for f in friends if f])):
        col_letter=get_column_letter(8+idx)
        ws_e.conditional_formatting.add(f'{col_letter}5:{col_letter}500', CellIsRule(operator='equal', formula=['"Y"'], fill=green_fill))
        ws_e.conditional_formatting.add(f'{col_letter}5:{col_letter}500', CellIsRule(operator='equal', formula=['"N"'], fill=gray_fill))

    ws_b=wb.create_sheet("4. Balances")
    for c,h in enumerate(["Friend","Total Paid","Total Share","Balance","Status","To Get/Pay"], start=1):
        cell=ws_b.cell(row=4,column=c,value=h)
        cell.fill=header_fill
        cell.font=header_font
    for idx,name in enumerate(friends):
        if not name: continue
        r=5+idx
        ws_b.cell(row=r,column=1,value=name).font=Font(bold=True)
        ws_b.cell(row=r,column=2,value=f"=SUMIF('3. Expenses - Mobile'!$E$5:$E$500,A{r},'3. Expenses - Mobile'!$F$5:$F$500)").number_format=money_fmt
        friend_col=get_column_letter(8+idx)
        formula=f"=SUMPRODUCT(('3. Expenses - Mobile'!$G$5:$G$500<>\"\")*('3. Expenses - Mobile'!${friend_col}$5:${friend_col}$500=\"Y\")*('3. Expenses - Mobile'!$F$5:$F$500/'3. Expenses - Mobile'!$G$5:$G$500))"
        ws_b.cell(row=r,column=3,value=formula).number_format=money_fmt
        ws_b.cell(row=r,column=4,value=f"=B{r}-C{r}").number_format=money_fmt
        ws_b.cell(row=r,column=5,value=f"=IF(D{r}>0.01,\"GETS BACK\",IF(D{r}<-0.01,\"OWES\",\"OK\"))")
        ws_b.cell(row=r,column=6,value=f"=ABS(D{r})").number_format=money_fmt
    ws_b.conditional_formatting.add('D5:D12', CellIsRule(operator='greaterThan', formula=['0.01'], fill=green_fill))
    ws_b.conditional_formatting.add('D5:D12', CellIsRule(operator='lessThan', formula=['-0.01'], fill=red_fill))

    ws_s=wb.create_sheet("5. Settlement")
    ws_s['A4']="Friend"
    ws_s['B4']="Balance"
    for c in [1,2]:
        ws_s.cell(row=4,column=c).fill=header_fill
        ws_s.cell(row=4,column=c).font=header_font
    for idx,name in enumerate(friends):
        if not name: continue
        r=5+idx
        ws_s.cell(row=r,column=1,value=f"='4. Balances'!A{5+idx}")
        ws_s.cell(row=r,column=2,value=f"='4. Balances'!D{5+idx}").number_format=money_fmt
    for t in range(10):
        ws_s.cell(row=4,column=3+t,value=f"After {t+1}").fill=sub_fill
    trans_start=16
    for c,h in enumerate(["Step","From","To","Amount","Note"], start=1):
        cell=ws_s.cell(row=trans_start,column=c,value=h)
        cell.fill=header_fill
        cell.font=header_font
    for t in range(10):
        tr=trans_start+1+t
        prev_letter=get_column_letter(2+t)
        ws_s.cell(row=tr,column=1,value=t+1)
        ws_s.cell(row=tr,column=2,value=f'=IFERROR(INDEX($A$5:$A$12,MATCH(TRUE,${prev_letter}$5:${prev_letter}$12<-0.01,0)),"")')
        ws_s.cell(row=tr,column=3,value=f'=IFERROR(INDEX($A$5:$A$12,MATCH(TRUE,${prev_letter}$5:${prev_letter}$12>0.01,0)),"")')
        ws_s.cell(row=tr,column=4,value=f'=IF(OR(B{tr}="",C{tr}=""),"",MIN(-INDEX(${prev_letter}$5:${prev_letter}$12,MATCH(B{tr},$A$5:$A$12,0)),INDEX(${prev_letter}$5:${prev_letter}$12,MATCH(C{tr},$A$5:$A$12,0))))').number_format=money_fmt
        ws_s.cell(row=tr,column=5,value=f'=IF(D{tr}="","",B{tr}&" → "&C{tr}&" : "&TEXT(D{tr},"#,##0"))')
    for idx in range(len([f for f in friends if f])):
        r=5+idx
        for t in range(10):
            tr=trans_start+1+t
            prev_letter=get_column_letter(2+t)
            ws_s.cell(row=r,column=3+t,value=f'={prev_letter}{r}+IF(B{tr}=$A{r},D{tr},0)-IF(C{tr}=$A{r},D{tr},0)').number_format=money_fmt

    output=io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"{trip['name']}_{trip['id']}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__=='__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
